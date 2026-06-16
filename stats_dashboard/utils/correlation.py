from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pingouin as pg
from scipy.stats import spearmanr
from statsmodels.stats.multitest import multipletests

from stats_dashboard.utils.loader import (
    get_network_category,
    is_random_graph,
    has_weight,
    get_graph_feature_columns,
    strip_z_prefix,
    METADATA_VARS,
    GRAPH_FEATURES,
    Z_GRAPH_FEATURES,
)

IMPULSIVITY_MEASURES = ["TOTAL", "NPLAN", "COG", "MOT", "MOT_V4", "COG_V1"]

COVARIATE = "School year"


def get_graph_features_for_network(network_type: str, task_label: str) -> list[str]:
    features = list(GRAPH_FEATURES)

    is_semantic_no_weight_no_random = (
        network_type.startswith("Semántico")
        and not has_weight(network_type)
        and not is_random_graph(network_type)
    )
    is_semantic_random_no_weight = (
        network_type.startswith("Semántico")
        and is_random_graph(network_type)
        and not has_weight(network_type)
    )
    is_any_random = is_random_graph(network_type)

    if is_semantic_no_weight_no_random:
        features = [f for f in features if f not in {"re", "pe", "l1", "l2", "l3", "cc"}]

    if is_semantic_random_no_weight:
        features = [f for f in features if f not in {"nodes", "edges", "re", "pe", "l1", "l2", "l3", "atd"}]

    if is_any_random:
        features = [f for f in features if f not in {"nodes", "edges"}]

    is_structural = get_network_category(network_type) == "Estructural"
    is_task6b_structural = task_label == "Task6-B" and is_structural
    if not is_task6b_structural and "cc" in features:
        features.remove("cc")

    return features


def resolve_column_name(col_clean: str, df: pd.DataFrame, network_type: str) -> str | None:
    if is_random_graph(network_type):
        z_col = f"z_{col_clean}"
        return z_col if z_col in df.columns else None
    return col_clean if col_clean in df.columns else None


def compute_simple_and_partial(
    df: pd.DataFrame, x_col: str, y_col: str, covar_col: str = COVARIATE
) -> dict[str, Any]:
    valid = df[[x_col, y_col, covar_col]].dropna()
    if len(valid) < 10:
        return {"rho_simple": None, "p_simple": None, "rho_partial": None, "p_partial": None}

    x = valid[x_col].values.astype(float)
    y = valid[y_col].values.astype(float)
    cov = valid[covar_col].values.astype(float)

    if np.nanvar(x) < 1e-12 or np.nanvar(y) < 1e-12:
        return {"rho_simple": None, "p_simple": None, "rho_partial": None, "p_partial": None}

    with np.errstate(all="ignore"):
        rho_s, p_s = spearmanr(x, y)

    rho_p, p_p = np.nan, np.nan
    try:
        pcorr_df = pd.DataFrame({"feat": x, "target": y, "cov": cov})
        pcorr = pg.partial_corr(
            data=pcorr_df, x="feat", y="target", covar="cov", method="spearman"
        )
        rho_p = pcorr["r"].values[0]
        pval_col = [c for c in pcorr.columns if c.startswith("p-") or c == "p_val"]
        if pval_col:
            p_p = pcorr[pval_col[0]].values[0]
        else:
            p_p = np.nan
    except Exception:
        rho_p, p_p = np.nan, np.nan

    return {
        "rho_simple": rho_s if not np.isnan(rho_s) else None,
        "p_simple": p_s if not np.isnan(p_s) else None,
        "rho_partial": rho_p if not np.isnan(rho_p) else None,
        "p_partial": p_p if not np.isnan(p_p) else None,
    }


def build_correlation_table(
    file_infos: list[dict],
    metadata_df: pd.DataFrame,
    network_type: str,
    task_label: str,
) -> pd.DataFrame:
    graph_features = get_graph_features_for_network(network_type, task_label)

    rows = []
    for imp_meas in IMPULSIVITY_MEASURES:
        for gf in graph_features:
            row = {"Impulsivity measures": imp_meas, "Graph features": gf}
            for finfo in file_infos:
                df_metric = finfo["df"]
                label = finfo["label"]
                merged = df_metric.merge(metadata_df, left_on="file", right_on="Cod", how="inner")

                actual_col = resolve_column_name(gf, merged, network_type)
                if actual_col is None or imp_meas not in merged.columns:
                    row[f"{label}_rho_simple"] = None
                    row[f"{label}_p_simple"] = None
                    row[f"{label}_rho_partial"] = None
                    row[f"{label}_p_partial"] = None
                    continue

                result = compute_simple_and_partial(
                    merged, actual_col, imp_meas, COVARIATE
                )
                row[f"{label}_rho_simple"] = result["rho_simple"]
                row[f"{label}_p_simple"] = result["p_simple"]
                row[f"{label}_rho_partial"] = result["rho_partial"]
                row[f"{label}_p_partial"] = result["p_partial"]

            rows.append(row)

    return pd.DataFrame(rows)


def apply_conditional_filtering(df: pd.DataFrame, file_labels: list[str]) -> pd.DataFrame:
    for label in file_labels:
        rho_s = df[f"{label}_rho_simple"]
        rho_p = df[f"{label}_rho_partial"]
        mask_s = rho_s.abs() < 0.1
        mask_p = rho_p.abs() < 0.1
        df.loc[mask_s, f"{label}_rho_simple"] = None
        df.loc[mask_s, f"{label}_p_simple"] = None
        df.loc[mask_p, f"{label}_rho_partial"] = None
        df.loc[mask_p, f"{label}_p_partial"] = None
    return df


def apply_corrections(
    df: pd.DataFrame, file_labels: list[str]
) -> tuple[pd.DataFrame, dict[str, float]]:
    correction_info = {}
    for label in file_labels:
        for imp in IMPULSIVITY_MEASURES:
            mask = (df["Impulsivity measures"] == imp) & df[f"{label}_p_partial"].notna()
            if mask.sum() < 2:
                continue
            p_vals = df.loc[mask, f"{label}_p_partial"].values

            _, p_bonf, _, _ = multipletests(p_vals, method="bonferroni")
            _, p_fdr, _, _ = multipletests(p_vals, method="fdr_bh")

            df.loc[mask, f"{label}_bonferroni"] = p_bonf
            df.loc[mask, f"{label}_fdr_pass"] = p_fdr < 0.05

            key = f"{imp}_{label}"
            correction_info[key] = {
                "bonferroni_alpha": 0.05 / len(p_vals),
                "fdr_q": 0.05,
                "n_tests": len(p_vals),
                "impulsivity": imp,
                "file_label": label,
            }

    return df, correction_info


def build_multiindex_columns(file_labels: list[str]) -> pd.MultiIndex:
    col_tuples = [("Impulsivity measures", ""), ("Graph features", "")]
    for label in file_labels:
        col_tuples.append((f"{label} - rho", ""))
        col_tuples.append((f"{label} - p", ""))
    for label in file_labels:
        col_tuples.append((f"{label} - Sy - rho", ""))
        col_tuples.append((f"{label} - Sy - p", ""))

    return pd.MultiIndex.from_tuples(col_tuples)


def flatten_df_for_display(
    df: pd.DataFrame, file_labels: list[str]
) -> pd.DataFrame:
    data = {}
    data[("Impulsivity measures", "")] = df["Impulsivity measures"]
    data[("Graph features", "")] = df["Graph features"]

    for label in file_labels:
        data[(label, "rho")] = df[f"{label}_rho_simple"]
        data[(label, "p-value")] = df[f"{label}_p_simple"]

    for label in file_labels:
        data[(f"{label} - Sy", "rho")] = df[f"{label}_rho_partial"]
        data[(f"{label} - Sy", "p-value")] = df[f"{label}_p_partial"]

    result = pd.DataFrame(data)
    return result


def style_table_html(
    df_display: pd.DataFrame,
    df_raw: pd.DataFrame,
    file_labels: list[str],
    correction_info: dict,
    zero_var_features: set[str] = frozenset(),
) -> str:
    html_parts = []

    first_alpha = None
    for cinfo in correction_info.values():
        first_alpha = cinfo["bonferroni_alpha"]
        break
    if first_alpha is not None:
        html_parts.append(
            f"<div style='margin-bottom:8px; font-size:0.9em;'>"
            f"Bonferroni α corregido: {first_alpha:.6f} | FDR q = 0.05"
            f"</div>"
        )

    n_files = len(file_labels)
    n_rows = len(df_display)

    imp_vals = df_display[("Impulsivity measures", "")].tolist()
    row_groups = []
    prev = None
    start = 0
    for i, v in enumerate(imp_vals):
        if v != prev:
            if prev is not None:
                row_groups.append({"imp": prev, "start": start, "end": i - 1})
            start = i
            prev = v
    if prev is not None:
        row_groups.append({"imp": prev, "start": start, "end": n_rows - 1})

    def _get_row_group(row_idx):
        for rg in row_groups:
            if rg["start"] <= row_idx <= rg["end"]:
                return rg
        return None

    col_groups = []
    for i in range(n_files):
        col_groups.append({"first": 2 + 2 * i, "last": 3 + 2 * i})
    offset = 2 + 2 * n_files
    for i in range(n_files):
        col_groups.append({"first": offset + 2 * i, "last": offset + 2 * i + 1})

    last_data_col = offset + 2 * n_files - 1

    def _col_borders(col_idx):
        left = ""
        right = ""
        for cg in col_groups:
            if col_idx == cg["first"]:
                left = "border-left:3px solid black;"
            if col_idx == cg["last"]:
                right = "border-right:3px solid black;"
        return left, right

    thick = "3px solid black"

    html_parts.append("<table border='1' cellpadding='6' cellspacing='0' style='border-collapse:collapse; font-size:0.85em;'>")

    header1 = ["<th colspan='2' style='background:#e0e0e0;'>Variables</th>"]
    header2 = ["<th style='background:#f5f5f5;'>Impulsivity measures</th>", "<th style='background:#f5f5f5;'>Graph features</th>"]

    for i, label in enumerate(file_labels):
        col_start = 2 + 2 * i
        l, r = _col_borders(col_start)
        extra = l + r
        header1.append(f"<th colspan='2' style='background:#d0e8f0; text-align:center;{extra}'>{label}</th>")
        for j in range(2):
            c = col_start + j
            l2, r2 = _col_borders(c)
            name = "rho" if j == 0 else "p-value"
            extra2 = l2 + r2 + "border-top:" + thick + ";"
            header2.append(f"<th style='background:#e8f4f8;{extra2}'>{name}</th>")

    for i, label in enumerate(file_labels):
        col_start = offset + 2 * i
        l, r = _col_borders(col_start)
        extra = l + r
        header1.append(f"<th colspan='2' style='background:#d0f0d0; text-align:center;{extra}'>{label} - Sy</th>")
        for j in range(2):
            c = col_start + j
            l2, r2 = _col_borders(c)
            name = "rho" if j == 0 else "p-value"
            extra2 = l2 + r2 + "border-top:" + thick + ";"
            header2.append(f"<th style='background:#e8f8e8;{extra2}'>{name}</th>")

    html_parts.append("<thead><tr>" + "".join(header1) + "</tr><tr>" + "".join(header2) + "</tr></thead>")
    html_parts.append("<tbody>")

    gray_data = "color:#c0c0c0;"
    has_zero_var_features = bool(zero_var_features)

    for idx in df_display.index:
        row_idx = df_display.index.get_loc(idx)
        rg = _get_row_group(row_idx)

        gf_name = str(df_display.iloc[idx][("Graph features", "")])
        is_zv = gf_name in zero_var_features

        border_top = f"border-top:{thick};" if rg and row_idx == rg["start"] else ""
        border_bot = f"border-bottom:{thick};" if rg and row_idx == rg["end"] else ""

        def _cell_style(col_idx, extra=""):
            l, r = _col_borders(col_idx)
            rg_left = "border-left:" + thick + ";" if col_idx == 0 else ""
            rg_right = "border-right:" + thick + ";" if col_idx == last_data_col else ""
            return border_top + border_bot + rg_left + rg_right + l + r + extra

        gf_style = _cell_style(1)
        if is_zv:
            gf_style += "color:#a0a0a0;font-style:italic;"
        data_extra = gray_data if is_zv else ""

        html_parts.append("<tr>")
        html_parts.append(
            f"<td style='{_cell_style(0)}font-weight:bold;'>"
            f"{df_display.iloc[idx][('Impulsivity measures', '')]}</td>"
        )
        html_parts.append(f"<td style='{gf_style}'>{gf_name}</td>")

        for col_offset, label in enumerate(file_labels):
            c = 2 + 2 * col_offset
            rho_s = df_display.iloc[idx][(label, "rho")]
            p_s = df_display.iloc[idx][(label, "p-value")]
            td_rho_s = f"<td style='{_cell_style(c)}{data_extra}'>{rho_s:.4f}</td>" if pd.notna(rho_s) else f"<td style='{_cell_style(c)}{data_extra}'></td>"
            td_p_s = f"<td style='{_cell_style(c + 1)}{data_extra}'>{p_s:.4f}</td>" if pd.notna(p_s) else f"<td style='{_cell_style(c + 1)}{data_extra}'></td>"
            html_parts.append(td_rho_s)
            html_parts.append(td_p_s)

        for col_offset, label in enumerate(file_labels):
            c = offset + 2 * col_offset
            rho_p = df_display.iloc[idx][(f"{label} - Sy", "rho")]
            p_p = df_display.iloc[idx][(f"{label} - Sy", "p-value")]

            if pd.notna(rho_p):
                color_style = "color:green; font-weight:bold;" if abs(rho_p) >= 0.1 else ""
                td_rho_p = f"<td style='{_cell_style(c)}{data_extra}{color_style}'>{rho_p:.4f}</td>"
            else:
                td_rho_p = f"<td style='{_cell_style(c)}{data_extra}'></td>"
            html_parts.append(td_rho_p)

            if pd.notna(p_p):
                imp = df_raw.iloc[idx]["Impulsivity measures"]
                label_key = f"{imp}_{label}"

                style = ""
                fdr_pass = False
                if label_key in correction_info:
                    cinfo = correction_info[label_key]
                    p_bonf_key = f"{label}_bonferroni"
                    fdr_key = f"{label}_fdr_pass"
                    if fdr_key in df_raw.columns and df_raw.iloc[idx][fdr_key]:
                        fdr_pass = True
                    if p_bonf_key in df_raw.columns and df_raw.iloc[idx][p_bonf_key] is not None and df_raw.iloc[idx][p_bonf_key] < 0.05:
                        style += "font-weight:bold;"

                if p_p > 0.05:
                    style += "color:gray;"

                td_p_p = f"<td style='{_cell_style(c + 1)}{data_extra}{style}'>{p_p:.4f}</td>"
                if fdr_pass:
                    td_p_p = td_p_p.replace("</td>", "*</td>")
            else:
                td_p_p = f"<td style='{_cell_style(c + 1)}{data_extra}'></td>"
            html_parts.append(td_p_p)

        html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    if has_zero_var_features:
        html_parts.append(
            "<div style='margin-top:8px; font-size:0.85em; color:#a0a0a0; font-style:italic;'>"
            "Las filas en gris itálico corresponden a variables con varianza cero; "
            "no se calcularon correlaciones y no se incluyeron en la corrección Bonferroni/FDR."
            "</div>"
        )
    return "\n".join(html_parts)


def export_to_excel(
    df_display: pd.DataFrame,
    df_raw: pd.DataFrame,
    file_labels: list[str],
    filename: str,
    correction_info: dict,
    zero_var_features: set[str] = frozenset(),
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output = io.BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Correlations"

    thin_side = Side(style="thin")
    thick_side = Side(style="thick")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    n_files = len(file_labels)
    n_rows = len(df_display)

    imp_vals = df_display[("Impulsivity measures", "")].tolist()
    row_groups = []
    prev = None
    start = 0
    for i, v in enumerate(imp_vals):
        if v != prev:
            if prev is not None:
                row_groups.append({"imp": prev, "start": start, "end": i - 1})
            start = i
            prev = v
    if prev is not None:
        row_groups.append({"imp": prev, "start": start, "end": n_rows - 1})

    def _get_row_group(row_idx):
        for rg in row_groups:
            if rg["start"] <= row_idx <= rg["end"]:
                return rg
        return None

    col_groups = []
    for i in range(n_files):
        col_groups.append({"first": 3 + 2 * i, "last": 4 + 2 * i})
    offset = 3 + 2 * n_files
    for i in range(n_files):
        col_groups.append({"first": offset + 2 * i, "last": offset + 2 * i + 1})
    last_data_col = offset + 2 * n_files - 1

    def _cell_border(row_idx, col_idx):
        top, bottom, left, right = thin_side, thin_side, thin_side, thin_side
        rg = _get_row_group(row_idx)
        if rg:
            if row_idx == rg["start"]:
                top = thick_side
            if row_idx == rg["end"]:
                bottom = thick_side
            if col_idx == 1:
                left = thick_side
            if col_idx == last_data_col:
                right = thick_side
        for cg in col_groups:
            if col_idx == cg["first"]:
                left = thick_side
            if col_idx == cg["last"]:
                right = thick_side
        return Border(left=left, right=right, top=top, bottom=bottom)

    def _set_border(cell, row_idx, col_idx):
        cell.border = _cell_border(row_idx, col_idx)

    header_fill_1 = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_fill_2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    simple_fill = PatternFill(start_color="D0E8F0", end_color="D0E8F0", fill_type="solid")
    simple_fill_2 = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    partial_fill = PatternFill(start_color="D0F0D0", end_color="D0F0D0", fill_type="solid")
    partial_fill_2 = PatternFill(start_color="E8F8E8", end_color="E8F8E8", fill_type="solid")
    green_font = Font(color="008000", bold=True)
    gray_font = Font(color="808080")
    bold_font = Font(bold=True)

    col = 1
    ws.cell(row=1, column=col, value="Variables").fill = header_fill_1
    ws.cell(row=1, column=col + 1).fill = header_fill_1
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    col += 2

    for label in file_labels:
        ws.cell(row=1, column=col, value=label).fill = simple_fill
        ws.cell(row=1, column=col + 1).fill = simple_fill
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        col += 2

    for label in file_labels:
        ws.cell(row=1, column=col, value=f"{label} - Sy").fill = partial_fill
        ws.cell(row=1, column=col + 1).fill = partial_fill
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        col += 2

    col = 1
    ws.cell(row=2, column=col, value="Impulsivity measures").fill = header_fill_2
    ws.cell(row=2, column=col + 1, value="Graph features").fill = header_fill_2
    ws.cell(row=2, column=col).border = thin_border
    ws.cell(row=2, column=col + 1).border = thin_border
    col += 2

    def _header_border(col_idx):
        top, bottom, left, right = thick_side, thin_side, thin_side, thin_side
        for cg in col_groups:
            if col_idx == cg["first"]:
                left = thick_side
            if col_idx == cg["last"]:
                right = thick_side
        return Border(left=left, right=right, top=top, bottom=bottom)

    for i, label in enumerate(file_labels):
        cg_cols = [3 + 2 * i, 4 + 2 * i]
        for c in cg_cols:
            cell = ws.cell(row=2, column=c, value=("rho" if c == cg_cols[0] else "p-value"))
            cell.fill = simple_fill_2
            cell.border = _header_border(c)
        col += 2

    for i, label in enumerate(file_labels):
        cg_cols = [offset + 2 * i, offset + 2 * i + 1]
        for c in cg_cols:
            cell = ws.cell(row=2, column=c, value=("rho" if c == cg_cols[0] else "p-value"))
            cell.fill = partial_fill_2
            cell.border = _header_border(c)
        col += 2

    gray_excel_font = Font(color="C0C0C0")
    gray_italic_font = Font(color="A0A0A0", italic=True)
    has_zero_var_features = bool(zero_var_features)

    for row_idx in range(len(df_display)):
        excel_row = row_idx + 3
        imp_val = df_display.iloc[row_idx][("Impulsivity measures", "")]
        gf_val = df_display.iloc[row_idx][("Graph features", "")]
        is_zv = has_zero_var_features and gf_val in zero_var_features

        c1 = ws.cell(row=excel_row, column=1, value=imp_val)
        _set_border(c1, row_idx, 1)
        c2 = ws.cell(row=excel_row, column=2, value=gf_val)
        _set_border(c2, row_idx, 2)

        if is_zv:
            c1.font = Font(bold=True, color="A0A0A0")
            c2.font = gray_italic_font
        else:
            c1.font = bold_font

        col = 3
        for label in file_labels:
            rho_s = df_display.iloc[row_idx][(label, "rho")]
            p_s = df_display.iloc[row_idx][(label, "p-value")]
            if pd.notna(rho_s):
                cell = ws.cell(row=excel_row, column=col, value=round(rho_s, 4))
            else:
                cell = ws.cell(row=excel_row, column=col, value="")
            _set_border(cell, row_idx, col)
            if is_zv:
                cell.font = gray_excel_font
            col += 1

            if pd.notna(p_s):
                cell = ws.cell(row=excel_row, column=col, value=round(p_s, 4))
            else:
                cell = ws.cell(row=excel_row, column=col, value="")
            _set_border(cell, row_idx, col)
            if is_zv:
                cell.font = gray_excel_font
            col += 1

        for label in file_labels:
            rho_p = df_display.iloc[row_idx][(f"{label} - Sy", "rho")]
            p_p = df_display.iloc[row_idx][(f"{label} - Sy", "p-value")]

            if pd.notna(rho_p):
                cell = ws.cell(row=excel_row, column=col, value=round(rho_p, 4))
                if abs(rho_p) >= 0.1 and not is_zv:
                    cell.font = green_font
            else:
                cell = ws.cell(row=excel_row, column=col, value="")
            _set_border(cell, row_idx, col)
            if is_zv:
                cell.font = gray_excel_font
            col += 1

            if pd.notna(p_p):
                cell = ws.cell(row=excel_row, column=col, value=round(p_p, 4))
                if p_p > 0.05 and not is_zv:
                    cell.font = gray_font
                if not is_zv:
                    imp = df_raw.iloc[row_idx]["Impulsivity measures"]
                    label_key = f"{imp}_{label}"
                    fdr_pass = False
                    if label_key in correction_info:
                        p_bonf_key = f"{label}_bonferroni"
                        fdr_key = f"{label}_fdr_pass"
                        if fdr_key in df_raw.columns and df_raw.iloc[row_idx][fdr_key]:
                            fdr_pass = True
                        if p_bonf_key in df_raw.columns and pd.notna(df_raw.iloc[row_idx].get(p_bonf_key, None)):
                            if df_raw.iloc[row_idx][p_bonf_key] < 0.05:
                                cell.font = Font(bold=True)
                    if fdr_pass:
                        cell.value = f"{p_p:.4f}*"
            else:
                cell = ws.cell(row=excel_row, column=col, value="")
            _set_border(cell, row_idx, col)
            if is_zv:
                cell.font = gray_excel_font
            col += 1

    for col_idx in range(1, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    wb.save(output)
    output.seek(0)
    return output.getvalue()
